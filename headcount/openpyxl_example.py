import time
from openpyxl import load_workbook

start = time.time() #Start Table creation timer
wb = load_workbook('data.xlsx')
end = time.time()
print ("load time is ", end-start)
# wb2 = load_workbook('data.ods')
#wb.get_sheet_names()
# wb2.get_sheet_names()

source = wb.get_sheet_by_name('raw data')
# print (wb.get_sheet_names())
# print (wb2.get_sheet_names())
#What's wrong with the ODS format?

# create set of unique Employee Numbers
# Create list of columns we need data from
start = time.time() #Start Table creation timer
table = []

criteria = ["Employee Num", "CC", "Company", "DOE/Project", "Total Hours"] #short for criteria




#l.update({100: {"project": 5 + l.get(100,{"project":0, "doe": 0})}})

