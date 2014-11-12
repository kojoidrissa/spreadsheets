from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import time

##new_stable: this is a branch from commit db6083 in Master. The NEXT commit, c8e3df, is where I changed the fullTable slice that messed up stuff.

##will this path work?: cd 'C:\Users\kidrissa\Documents\Monthly Headcount Schedule\July 2013 Headcount'
    ##Yes, but only from the shell. I'll need to bring in OS Module to make this work in code
    ##OS Module may also help as a stopgap for the other modules in funcationalSheets.py; but setup.py is the REAL solution

#starting timer; basic performance profiling
start0 = time.time() #Start Loading Timer

wb = load_workbook(filename = r'data.xlsx')
source = wb.get_sheet_by_name('raw data')
end = time.time() #End Timer

durLoad = end - start0 #duration to Load

"""
    Beginning of New 'worksheet_to_table' Function                            
    OpenPyXl Worksheet Object --> Nested list of lists                        
    Purpose: put workbook values in a format Python can work with more easily 
"""
#Building a list of lists; each internal list represents a row of data; should this be a function? Probably.
start = time.time() #Start Table creation timer
table = []

#This will be less fragile if I take the following advice from Glen:
    #do this by column header/name instead of index
    #include code that will throw a VISIBLE exception if a needed column is missing
for row in range(len(source.rows)):
    r =[]
    ref = (3, 2, 0, 10, 4, 17, 21) 
    for col in ref: #Original "in" argument was 'range(len(source.columns))'
        r.append(source.cell(row = row, column = col).value)
    table.append(r)
end = time.time() #End Timer
durTable = end - start
"""End of 'worksheet_to_table' function"""

#Only for viewing profiling results
print(("Loading time for", source, ": ", durLoad))
print(("Time to create 'Table' from ",source, "for", len(source.rows), "rows and ", len(source.columns), "columns: ", durTable))


#Creating a spreadsheet in memory; Writing results to it (in memory)

start = time.time() #Start timer for creating 'Target' workbook
target = Workbook()
dest_filename = r'hdcntsum.xlsx'

# regarding a question about relative paths
# both r'Downloads\hdcntsum.xlsx' and r'..\hdcntsum.xlsx worked
# haven't tested networked drives yet
# my theory is 
#     r'..\..\..\M:\Dbsteam\BUDGET\Jackie\MNTH_RPT\2013\June 2013\Headcount Misc\hdcntsum.xlsx' SHOULD work
# Maybe not: r'\hdcntsum.xlsx' saved the file to the ROOT drive; C:
# OS Module may be the answer here as well
       

"""
    Function: create_keylist
        list of lists --> list of lists 
    takes 'table' (a list of lists) from worksheet_to_table
    returns a nested list of keys; this should probably be a list of tuples

"""
#Creating list of keys
#Each key is (as of 2013-07-16)a LIST made up of [Company Number, CC, EmpNum]. It SHOULD be
#a TUPLE made up of (Company Number, CC, EmpNum). But I was having trouble with Tuples
#I'll fix that in a refactor. Tuples are better for keys. Immutibilty FTW!
start = time.time() #Start timer for creating 'Keylist' workbook
keylist = []

for r in range(len(table)): #I need to change the range to (1, len(table)), to get rid of the header key
    newkey = table[r][:3] #Original code was 'newkey = tuple(table[r][:3])'
    if newkey not in keylist:
        keylist.append(newkey)
end = time.time()  #End timer for creating 'Keylist' 
durKeylist = end - start #durATION for Keylist
"""end create_keylist function"""


"""
    Function: hourtable
        keylist(list of lists), table(list of lists) --> hourtable(list of lists)
    Comparing keylist items to table rows and calculating hour totals
"""
"""
    THIS IS THE FUNCTION THAT NEEDS TO BE UPDATED TO SHOW TIME SPLIT into
    three CATEGORIES (Proj/non-Proj/ATO) INSTEAD OF two
"""

#using 'r' as shorthand for 'row', to avoid namespace confusion
start = time.time() #start Hourtable timer
hourtable = []
#'hourtable' will be a list of lists; each sublist is composed of a tuple and two ints; 
#Index 0 of each list is the composite key; Index 1 is DOE; Index 2 is Project;
#DOE and Project are the hourly totals FOR that key
for key in keylist:
    doe = 0 #Counter for DOE Hours
    project = 0 #Counter for Project hours
    newrow =[]
    for r in range(len(table)):
        if key == table[r][:3]: #Tuple version: 'if key == tuple(table[r][:3])'
            if table[r][5] == 'DOE': ###CHANGE TO 'NON-PROJECT' IN UPDATE
                doe = doe + table[r][6]
            elif table[r][5] == 'Project':
                project = project + table[r][6]
    newrow.extend(key) #add key values to the new row
    newrow.extend([doe, project]) #add total DOE & Project hours to new row
    hourtable.append(newrow) #add new row to table; Was hourtable.append([key, doe, project])
end = time.time()  #End timer for creating 'Hourlist'
durHourlist = end - start
"""end hourtable function"""


"""
    Function: finaltable
        keylist(list of lists), table(list of lists), hourtable(list of lists) --> finaltable(list of lists)
    Create final output table by matching rows b/w table and hourtable with keylist, then combining
    those matched rows into the final output form
"""
start = time.time() #start finaltable creation timer

finaltable =[]
for k in range(len(keylist)):
    finalrow = []
    for r in range(len(table)):
        if keylist[k][:3] == table[r][:3]: 
            finalrow = table[r][:5] + hourtable[k][-2:]
    finaltable.append(finalrow)

end = time.time()  #End 'finaltable' timer
durFinalTable = end - start
"""end finaltable function"""

#Write final workbook to memory and save to file

start = time.time() #start Target final spreadsheet write to memory timer

ws1 = target.create_sheet(0)
ws1.title = "Monthly Headcount Summary"
for row in finaltable:
    rowIn = finaltable.index(row)
    for col in range(len(finaltable[0])): #changed from "range(len(finaltable[0]))"
        #colIn = row.index(col) #colIn replaces col as the Column Indexes below
        ws1.cell(row = rowIn, column = col).value = finaltable[rowIn][col]
        
end = time.time() #End Target final spreadsheet write to memory timer
durFinalTableMem = end - start

#Writing that worksheet to a file
start = time.time() #start Target final spreadsheet write to file timer

target.save(dest_filename)

end0 = time.time() #End Target final spreadsheet write to file timer
durFinalTableFile = end0 - start

durTotal = end0 - start0 

#Printing my timer variables
print(("Loading time for", source, " :", durLoad))
print((len(source.rows),"Rows; ", len(source.columns), "Columns"))
print("durTable", durTable)
print(len(table),"Rows; ", len(table[0]), "Columns")
print("durKeylist", durKeylist)
print("durHourlist", durHourlist)
print("durFinalTable", durFinalTable)
print("durFinalTableMem", durFinalTableMem)
print("Writing time for", dest_filename, " :", durFinalTableFile)
print(len(finaltable),"Rows; ", len(finaltable[0]), "Columns")
print("durTotal", durTotal)